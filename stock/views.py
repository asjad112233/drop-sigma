import csv
import io
import json
import uuid
import requests as _req
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.db import transaction

from stores.models import Store
from orders.models import Order
from .models import StockProduct, StockVariant, StockEntry, StockAuditLog, StockOrderAssignment, StockAutoRule


def _actor(request):
    if request.user.is_authenticated:
        name = request.user.get_full_name() or request.user.username
        return name
    return "Admin"


def _variant_data(variant):
    entry = getattr(variant, "entry", None)
    return {
        "id": variant.id,
        "color": variant.color,
        "size": variant.size,
        "sku": variant.sku,
        "quantity": entry.quantity if entry else 0,
        "reserved": entry.reserved if entry else 0,
        "available": entry.available() if entry else 0,
    }


def _product_data(product, include_variants=True):
    variants = list(product.variants.all()) if include_variants else []
    total_qty = 0
    total_available = 0
    for v in variants:
        e = getattr(v, "entry", None)
        if e:
            total_qty += e.quantity
            total_available += e.available()
    return {
        "id": product.id,
        "product_id": product.product_id,
        "product_name": product.product_name,
        "image_url": product.image_url or "",
        "store_id": product.store_id,
        "store_name": product.store.name,
        "is_active": product.is_active,
        "synced_at": product.synced_at.isoformat(),
        "variants": [_variant_data(v) for v in variants] if include_variants else [],
        "total_qty": total_qty,
        "total_available": total_available,
        "variant_count": len(variants),
    }


@login_required
@require_http_methods(["GET"])
def stock_dashboard_api(request):
    store_id = request.GET.get("store_id")
    qs = StockProduct.objects.select_related("store").prefetch_related("variants__entry")
    if store_id:
        qs = qs.filter(store_id=store_id)

    products = list(qs)
    all_entries = StockEntry.objects.filter(variant__product__in=products)
    total_qty = sum(e.quantity for e in all_entries)
    low_stock = sum(1 for e in all_entries if 0 < e.quantity <= 5)
    out_of_stock = sum(1 for e in all_entries if e.quantity == 0)
    total_variants = sum(p.variant_count for p in [])
    tv = StockVariant.objects.filter(product__in=products).count()

    return JsonResponse({
        "success": True,
        "stats": {
            "total_products": len(products),
            "total_variants": tv,
            "total_qty": total_qty,
            "low_stock": low_stock,
            "out_of_stock": out_of_stock,
        },
        "products": [_product_data(p) for p in products],
    })


@login_required
@require_http_methods(["POST"])
def stock_sync_api(request):
    data = json.loads(request.body)
    store_id = data.get("store_id")
    if not store_id:
        return JsonResponse({"success": False, "message": "store_id required"}, status=400)

    try:
        store = Store.objects.get(id=store_id)
    except Store.DoesNotExist:
        return JsonResponse({"success": False, "message": "Store not found"}, status=404)

    orders = Order.objects.filter(store=store).exclude(product_id__isnull=True).exclude(product_id="")
    synced = 0
    for order in orders:
        pid = str(order.product_id).strip()
        pname = (order.product_name or pid).strip()
        if not pid:
            continue
        product, _ = StockProduct.objects.get_or_create(
            store=store, product_id=pid,
            defaults={"product_name": pname},
        )

        # Try to get variants from raw order data
        raw = order.raw_data or {}
        line_items = raw.get("line_items", [])
        found_variant = False
        for item in line_items:
            color, size, sku = "", "", ""
            for prop in item.get("properties", []):
                k = str(prop.get("name", "")).lower()
                v = str(prop.get("value", ""))
                if k in ("color", "colour"):
                    color = v
                elif k == "size":
                    size = v
            sku = item.get("sku", "") or ""
            variant, _ = StockVariant.objects.get_or_create(
                product=product, color=color, size=size, defaults={"sku": sku}
            )
            StockEntry.objects.get_or_create(variant=variant)
            found_variant = True

        if not found_variant:
            variant, _ = StockVariant.objects.get_or_create(
                product=product, color="", size="", defaults={"sku": ""}
            )
            StockEntry.objects.get_or_create(variant=variant)

        synced += 1

    return JsonResponse({"success": True, "synced": synced})


@login_required
@require_http_methods(["GET", "POST"])
def stock_entry_api(request):
    if request.method == "GET":
        variant_id = request.GET.get("variant_id")
        try:
            variant = StockVariant.objects.select_related("product__store").get(id=variant_id)
        except StockVariant.DoesNotExist:
            return JsonResponse({"success": False, "message": "Variant not found"}, status=404)
        StockEntry.objects.get_or_create(variant=variant)
        return JsonResponse({"success": True, "variant": _variant_data(variant)})

    data = json.loads(request.body)
    variant_id = data.get("variant_id")
    new_qty = data.get("quantity")
    note = data.get("note", "")
    action = data.get("action", "adjust")

    try:
        variant = StockVariant.objects.get(id=variant_id)
    except StockVariant.DoesNotExist:
        return JsonResponse({"success": False, "message": "Variant not found"}, status=404)

    entry, _ = StockEntry.objects.get_or_create(variant=variant)
    qty_before = entry.quantity
    if new_qty is not None:
        entry.quantity = max(0, int(new_qty))
    entry.updated_by = request.user
    entry.save()

    StockAuditLog.objects.create(
        variant=variant, action=action,
        qty_before=qty_before, qty_after=entry.quantity,
        actor=_actor(request), note=note,
    )
    return JsonResponse({"success": True, "quantity": entry.quantity, "available": entry.available()})


@login_required
@require_http_methods(["POST"])
def stock_add_product_api(request):
    data = json.loads(request.body)
    store_id = data.get("store_id")
    product_name = data.get("product_name", "").strip()
    color = data.get("color", "").strip()
    size = data.get("size", "").strip()
    sku = data.get("sku", "").strip()
    quantity = max(0, int(data.get("quantity", 0) or 0))

    if not store_id or not product_name:
        return JsonResponse({"success": False, "message": "store_id and product_name required"}, status=400)

    try:
        store = Store.objects.get(id=store_id)
    except Store.DoesNotExist:
        return JsonResponse({"success": False, "message": "Store not found"}, status=404)

    product_id = data.get("product_id") or f"manual-{uuid.uuid4().hex[:8]}"

    with transaction.atomic():
        product, _ = StockProduct.objects.get_or_create(
            store=store, product_id=product_id,
            defaults={"product_name": product_name},
        )
        variant, _ = StockVariant.objects.get_or_create(
            product=product, color=color, size=size, defaults={"sku": sku}
        )
        entry, _ = StockEntry.objects.get_or_create(variant=variant)
        qty_before = entry.quantity
        entry.quantity = max(0, entry.quantity + quantity)
        entry.updated_by = request.user
        entry.save()

        StockAuditLog.objects.create(
            variant=variant, action="add",
            qty_before=qty_before, qty_after=entry.quantity,
            actor=_actor(request), note=data.get("note", "Manual add"),
        )

    return JsonResponse({"success": True, "product_id": product.id, "variant_id": variant.id})


def _parse_upload_rows(file_obj, filename):
    """Parse CSV or XLSX file into list of dicts. Returns (rows, error_str)."""
    name = filename.lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip().lower() if h else "" for h in next(rows_iter, [])]
        result = []
        for row in rows_iter:
            result.append({headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(headers))})
        return result, None
    else:
        text = file_obj.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [{k.strip().lower(): (v or "").strip() for k, v in row.items()} for row in reader], None


@login_required
@require_http_methods(["POST"])
def stock_bulk_upload_api(request):
    store_id = request.POST.get("store_id")
    upload_file = request.FILES.get("file")
    if not store_id or not upload_file:
        return JsonResponse({"success": False, "message": "store_id and file required"}, status=400)

    try:
        store = Store.objects.get(id=store_id)
    except Store.DoesNotExist:
        return JsonResponse({"success": False, "message": "Store not found"}, status=404)

    data_rows, err = _parse_upload_rows(upload_file, upload_file.name)
    if err:
        return JsonResponse({"success": False, "message": err}, status=400)

    rows_ok, rows_err = 0, 0
    errors = []

    with transaction.atomic():
        for i, row in enumerate(data_rows, start=2):
            try:
                pname = (row.get("product_name") or "").strip()
                color = (row.get("color") or "").strip()
                size = (row.get("size") or "").strip()
                sku = (row.get("sku") or "").strip()
                qty = max(0, int(float(row.get("quantity") or 0)))
                if not pname:
                    raise ValueError("product_name is empty")

                pid = (row.get("product_id") or "").strip() or f"csv-{pname[:30].lower().replace(' ', '-')}"
                product, _ = StockProduct.objects.get_or_create(
                    store=store, product_id=pid, defaults={"product_name": pname}
                )
                variant, _ = StockVariant.objects.get_or_create(
                    product=product, color=color, size=size, defaults={"sku": sku}
                )
                entry, _ = StockEntry.objects.get_or_create(variant=variant)
                qty_before = entry.quantity
                entry.quantity = qty
                entry.updated_by = request.user
                entry.save()

                StockAuditLog.objects.create(
                    variant=variant, action="add",
                    qty_before=qty_before, qty_after=qty,
                    actor=_actor(request), note="Bulk upload",
                )
                rows_ok += 1
            except Exception as e:
                rows_err += 1
                errors.append(f"Row {i}: {e}")

    return JsonResponse({"success": True, "imported": rows_ok, "errors": rows_err, "error_details": errors[:20]})


@login_required
@require_http_methods(["POST"])
def stock_deduct_api(request):
    data = json.loads(request.body)
    variant_id = data.get("variant_id")
    qty = int(data.get("quantity", 1))
    order_id = data.get("order_id")

    try:
        variant = StockVariant.objects.get(id=variant_id)
    except StockVariant.DoesNotExist:
        return JsonResponse({"success": False, "message": "Variant not found"}, status=404)

    entry, _ = StockEntry.objects.get_or_create(variant=variant)
    qty_before = entry.quantity
    entry.quantity = max(0, entry.quantity - qty)
    entry.updated_by = request.user
    entry.save()

    order = None
    if order_id:
        try:
            order = Order.objects.get(id=order_id)
        except Order.DoesNotExist:
            pass

    StockAuditLog.objects.create(
        variant=variant, order=order, action="deduct",
        qty_before=qty_before, qty_after=entry.quantity,
        actor=_actor(request), note=data.get("note", ""),
    )
    return JsonResponse({"success": True, "quantity": entry.quantity, "available": entry.available()})


@login_required
@require_http_methods(["GET"])
def stock_audit_api(request):
    variant_id = request.GET.get("variant_id")
    product_id = request.GET.get("product_id")
    store_id = request.GET.get("store_id")
    limit = min(int(request.GET.get("limit", 100)), 500)

    qs = StockAuditLog.objects.select_related("variant__product__store", "order")

    if variant_id:
        qs = qs.filter(variant_id=variant_id)
    elif product_id:
        qs = qs.filter(variant__product_id=product_id)
    elif store_id:
        qs = qs.filter(variant__product__store_id=store_id)

    logs = []
    for log in qs[:limit]:
        logs.append({
            "id": log.id,
            "action": log.action,
            "action_label": log.get_action_display(),
            "variant": str(log.variant),
            "product_name": log.variant.product.product_name,
            "color": log.variant.color,
            "size": log.variant.size,
            "qty_before": log.qty_before,
            "qty_after": log.qty_after,
            "actor": log.actor,
            "note": log.note,
            "order_id": log.order_id,
            "created_at": log.created_at.isoformat(),
        })

    return JsonResponse({"success": True, "logs": logs})


@login_required
@require_http_methods(["GET"])
def stock_fetch_store_products_api(request):
    """Fetch live product list from WooCommerce/Shopify store API."""
    store_id = request.GET.get("store_id")
    if not store_id:
        return JsonResponse({"success": False, "message": "store_id required"}, status=400)

    try:
        store = Store.objects.get(id=store_id)
    except Store.DoesNotExist:
        return JsonResponse({"success": False, "message": "Store not found"}, status=404)

    products = []

    try:
        if store.platform == "woocommerce":
            base = store.store_url.rstrip("/")
            auth = (store.api_key, store.api_secret)
            page = 1
            while True:
                r = _req.get(
                    f"{base}/wp-json/wc/v3/products",
                    auth=auth,
                    params={"per_page": 100, "page": page, "status": "publish"},
                    timeout=15,
                )
                if r.status_code != 200:
                    break
                batch = r.json()
                if not batch:
                    break
                for p in batch:
                    variants = []
                    if p.get("type") == "variable":
                        vr = _req.get(
                            f"{base}/wp-json/wc/v3/products/{p['id']}/variations",
                            auth=auth, params={"per_page": 100}, timeout=15,
                        )
                        if vr.status_code == 200:
                            for v in vr.json():
                                color, size = "", ""
                                for attr in v.get("attributes", []):
                                    name = attr.get("name", "").lower()
                                    val = attr.get("option", "")
                                    if name in ("color", "colour"):
                                        color = val
                                    elif name == "size":
                                        size = val
                                variants.append({
                                    "variant_id": str(v["id"]),
                                    "color": color,
                                    "size": size,
                                    "sku": v.get("sku", ""),
                                    "stock_quantity": v.get("stock_quantity") or 0,
                                })
                    else:
                        variants.append({
                            "variant_id": str(p["id"]),
                            "color": "", "size": "",
                            "sku": p.get("sku", ""),
                            "stock_quantity": p.get("stock_quantity") or 0,
                        })
                    images = p.get("images", [])
                    image_url = images[0].get("src", "") if images else ""
                    products.append({
                        "product_id": str(p["id"]),
                        "product_name": p.get("name", ""),
                        "image_url": image_url,
                        "variants": variants,
                        "already_imported": StockProduct.objects.filter(
                            store=store, product_id=str(p["id"])
                        ).exists(),
                    })
                if len(batch) < 100:
                    break
                page += 1

        elif store.platform == "shopify":
            base = store.store_url.rstrip("/")
            headers = {"X-Shopify-Access-Token": store.access_token} if store.access_token else {}
            r = _req.get(
                f"{base}/admin/api/2024-01/products.json",
                headers=headers, params={"limit": 250, "status": "active"}, timeout=15,
            )
            if r.status_code == 200:
                for p in r.json().get("products", []):
                    variants = []
                    for v in p.get("variants", []):
                        variants.append({
                            "variant_id": str(v["id"]),
                            "color": v.get("option1", "") or "",
                            "size": v.get("option2", "") or "",
                            "sku": v.get("sku", ""),
                            "stock_quantity": v.get("inventory_quantity") or 0,
                        })
                    images = p.get("images", [])
                    image_url = images[0].get("src", "") if images else ""
                    products.append({
                        "product_id": str(p["id"]),
                        "product_name": p.get("title", ""),
                        "image_url": image_url,
                        "variants": variants,
                        "already_imported": StockProduct.objects.filter(
                            store=store, product_id=str(p["id"])
                        ).exists(),
                    })
        else:
            return JsonResponse({"success": False, "message": "Platform not supported"}, status=400)

    except Exception as e:
        return JsonResponse({"success": False, "message": f"Store API error: {str(e)[:200]}"}, status=502)

    return JsonResponse({"success": True, "products": products, "total": len(products)})


@login_required
@require_http_methods(["POST"])
def stock_import_products_api(request):
    """Import selected products from store API into stock."""
    data = json.loads(request.body)
    store_id = data.get("store_id")
    selected = data.get("products", [])  # list of {product_id, product_name, variants:[{variant_id,color,size,sku}]}

    if not store_id or not selected:
        return JsonResponse({"success": False, "message": "store_id and products required"}, status=400)

    try:
        store = Store.objects.get(id=store_id)
    except Store.DoesNotExist:
        return JsonResponse({"success": False, "message": "Store not found"}, status=404)

    imported = 0
    with transaction.atomic():
        for item in selected:
            pid = str(item.get("product_id", "")).strip()
            pname = item.get("product_name", pid).strip()
            image_url = item.get("image_url", "") or ""
            if not pid:
                continue
            product, created = StockProduct.objects.get_or_create(
                store=store, product_id=pid, defaults={"product_name": pname, "image_url": image_url}
            )
            if not created and image_url and not product.image_url:
                product.image_url = image_url
                product.save(update_fields=["image_url"])
            for v in item.get("variants", []):
                color = (v.get("color") or "").strip()
                size = (v.get("size") or "").strip()
                sku = (v.get("sku") or "").strip()
                variant, _ = StockVariant.objects.get_or_create(
                    product=product, color=color, size=size, defaults={"sku": sku}
                )
                entry, created = StockEntry.objects.get_or_create(variant=variant)
                if created:
                    StockAuditLog.objects.create(
                        variant=variant, action="sync",
                        qty_before=0, qty_after=0,
                        actor=_actor(request), note="Imported from store",
                    )
            imported += 1

    return JsonResponse({"success": True, "imported": imported})


@login_required
@require_http_methods(["GET"])
def stock_export_api(request):
    """Export stock as CSV or XLSX download."""
    store_id = request.GET.get("store_id")
    fmt = request.GET.get("format", "csv").lower()
    qs = StockVariant.objects.select_related("product__store", "entry")
    if store_id:
        qs = qs.filter(product__store_id=store_id)

    headers = ["Product ID", "Product Name", "Store", "Color", "Size", "SKU", "Quantity", "Reserved", "Available"]
    rows = []
    for v in qs:
        e = getattr(v, "entry", None)
        rows.append([
            v.product.product_id,
            v.product.product_name,
            v.product.store.name,
            v.color,
            v.size,
            v.sku,
            e.quantity if e else 0,
            e.reserved if e else 0,
            e.available() if e else 0,
        ])

    if fmt == "xlsx":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock"

        # Header row styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(fill_type="solid", fgColor="7C3AED")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto column widths
        col_widths = [max(len(str(h)), max((len(str(r[i])) for r in rows), default=0)) + 4
                      for i, h in enumerate(headers)]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(width, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="stock_export.xlsx"'
        return response

    # Default: CSV
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="stock_export.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response


@login_required
@require_http_methods(["POST"])
def stock_bulk_update_api(request):
    """Bulk update quantity for multiple variants."""
    data = json.loads(request.body)
    updates = data.get("updates", [])  # [{variant_id, quantity, note}]

    if not updates:
        return JsonResponse({"success": False, "message": "No updates provided"}, status=400)

    updated = 0
    with transaction.atomic():
        for item in updates:
            variant_id = item.get("variant_id")
            new_qty = item.get("quantity")
            note = item.get("note", "Bulk update")
            if variant_id is None or new_qty is None:
                continue
            try:
                variant = StockVariant.objects.get(id=variant_id)
            except StockVariant.DoesNotExist:
                continue
            entry, _ = StockEntry.objects.get_or_create(variant=variant)
            qty_before = entry.quantity
            entry.quantity = max(0, int(new_qty))
            entry.updated_by = request.user
            entry.save()
            StockAuditLog.objects.create(
                variant=variant, action="adjust",
                qty_before=qty_before, qty_after=entry.quantity,
                actor=_actor(request), note=note,
            )
            updated += 1

    return JsonResponse({"success": True, "updated": updated})


# ─── Assign Stock to Order ────────────────────────────────────────────────────

from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def stock_assign_order_api(request):
    """POST: assign stock variants to order line items (+ optional permanent rule)."""
    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=405)

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    order_id   = data.get("order_id")
    assignments = data.get("assignments", [])  # [{product_id, variant_id, quantity}]
    permanent  = bool(data.get("permanent", False))
    store_id   = data.get("store_id")

    try:
        order = Order.objects.get(id=order_id)
    except Order.DoesNotExist:
        return JsonResponse({"error": "Order not found"}, status=404)

    results = []
    with transaction.atomic():
        for a in assignments:
            variant_id = a.get("variant_id")
            product_id = str(a.get("product_id", ""))
            quantity   = max(1, int(a.get("quantity", 1)))

            try:
                variant = StockVariant.objects.select_related("product").get(id=variant_id)
            except StockVariant.DoesNotExist:
                results.append({"product_id": product_id, "ok": False, "error": "Variant not found"})
                continue

            entry, _ = StockEntry.objects.get_or_create(variant=variant)
            qty_before = entry.quantity
            entry.quantity = max(0, entry.quantity - quantity)
            entry.save()

            StockOrderAssignment.objects.update_or_create(
                order=order, product_id=product_id,
                defaults={"variant": variant, "quantity": quantity},
            )

            StockAuditLog.objects.create(
                variant=variant, order=order, action="deduct",
                qty_before=qty_before, qty_after=entry.quantity,
                actor=_actor(request),
                note=f"Assigned to Order #{order.external_order_id or order.id}",
            )

            if permanent and store_id:
                try:
                    store = Store.objects.get(id=store_id)
                    StockAutoRule.objects.update_or_create(
                        store=store, product_id=product_id,
                        defaults={"variant": variant},
                    )
                except Store.DoesNotExist:
                    pass

            results.append({"product_id": product_id, "variant_id": variant_id, "ok": True})

    return JsonResponse({"success": True, "results": results})


@csrf_exempt
def stock_order_assignments_api(request):
    """GET: return stock assignment status for orders in a store.
       DELETE ?rule_id=X : remove a permanent auto-rule."""

    if request.method == "DELETE":
        rule_id = request.GET.get("rule_id")
        StockAutoRule.objects.filter(id=rule_id).delete()
        return JsonResponse({"success": True})

    store_id = request.GET.get("store_id")
    qs = StockOrderAssignment.objects.select_related(
        "variant__product", "variant__entry", "order"
    )
    if store_id:
        qs = qs.filter(order__store_id=store_id)

    assignments = {}
    for a in qs:
        oid = str(a.order_id)
        if oid not in assignments:
            assignments[oid] = {"assigned": True, "items": []}
        stock_left = 0
        try:
            stock_left = a.variant.entry.quantity
        except Exception:
            pass
        assignments[oid]["items"].append({
            "product_id": a.product_id,
            "variant_id": a.variant_id,
            "variant_name": str(a.variant),
            "sku": a.variant.sku,
            "quantity": a.quantity,
            "stock_left": stock_left,
        })

    auto_rules = {}
    if store_id:
        for r in StockAutoRule.objects.filter(store_id=store_id).select_related("variant"):
            auto_rules[r.product_id] = {
                "rule_id": r.id,
                "variant_id": r.variant_id,
                "variant_name": str(r.variant),
                "sku": r.variant.sku,
            }

    return JsonResponse({"success": True, "assignments": assignments, "auto_rules": auto_rules})


# ─── Stock Orders Management ─────────────────────────────────────────────────

@csrf_exempt
def stock_orders_api(request):
    """GET: all orders that have stock assigned + their details.
       POST action=add_tracking: save tracking number to order."""

    if request.method == "POST":
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        action = data.get("action")
        if action == "add_tracking":
            order_id = data.get("order_id")
            tracking_number  = data.get("tracking_number", "").strip()
            tracking_company = data.get("tracking_company", "").strip()
            tracking_url     = data.get("tracking_url", "").strip()
            try:
                order = Order.objects.get(id=order_id)
                order.tracking_number  = tracking_number
                order.tracking_company = tracking_company
                if tracking_url:
                    order.tracking_url = tracking_url
                if tracking_number:
                    order.tracking_status     = "shipped"
                    order.fulfillment_status  = "shipped"
                else:
                    order.tracking_status = "pending"
                order.save(update_fields=["tracking_number", "tracking_company", "tracking_url", "tracking_status", "fulfillment_status"])
                return JsonResponse({"success": True})
            except Order.DoesNotExist:
                return JsonResponse({"error": "Order not found"}, status=404)

        if action == "remove_assignment":
            assignment_id = data.get("assignment_id")
            try:
                a = StockOrderAssignment.objects.select_related("variant__entry").get(id=assignment_id)
                # Restore stock
                entry, _ = StockEntry.objects.get_or_create(variant=a.variant)
                qty_before = entry.quantity
                entry.quantity += a.quantity
                entry.save()
                StockAuditLog.objects.create(
                    variant=a.variant, action="restore",
                    qty_before=qty_before, qty_after=entry.quantity,
                    actor=_actor(request),
                    note=f"Assignment removed from Order #{a.order.external_order_id or a.order.id}",
                )
                a.delete()
                return JsonResponse({"success": True})
            except StockOrderAssignment.DoesNotExist:
                return JsonResponse({"error": "Assignment not found"}, status=404)

        return JsonResponse({"error": "Unknown action"}, status=400)

    # ── GET ──
    store_id = request.GET.get("store_id")
    qs = StockOrderAssignment.objects.select_related(
        "order__store", "order__assigned_vendor",
        "variant__product", "variant__entry",
    )
    if store_id:
        qs = qs.filter(order__store_id=store_id)

    # Group by order
    order_map = {}
    for a in qs:
        oid = a.order_id
        if oid not in order_map:
            o = a.order
            has_tracking = bool(o.tracking_number and o.tracking_number.lower() not in ("", "pending", "no tracking"))
            order_map[oid] = {
                "assignment_id": a.id,
                "order_id":      o.id,
                "order_number":  o.external_order_id or str(o.id),
                "customer_name": o.customer_name or "—",
                "customer_email": o.customer_email or "",
                "store_name":    o.store.name,
                "payment_status":     o.payment_status or "",
                "fulfillment_status": o.fulfillment_status or "",
                "total_price":   str(o.total_price or ""),
                "currency":      o.currency or "",
                "tracking_number":  o.tracking_number or "",
                "tracking_company": o.tracking_company or "",
                "tracking_url":     getattr(o, "tracking_url", "") or "",
                "tracking_status":  o.tracking_status or "pending",
                "has_tracking":  has_tracking,
                "vendor_name":   o.assigned_vendor.name if o.assigned_vendor else "",
                "created_at":    o.created_at.isoformat() if o.created_at else "",
                "items": [],
            }
        stock_left = 0
        try:
            stock_left = a.variant.entry.quantity
        except Exception:
            pass

        # Check if auto-rule exists for this product
        has_auto_rule = StockAutoRule.objects.filter(
            store=a.order.store, product_id=a.product_id
        ).exists()

        order_map[oid]["items"].append({
            "assignment_id": a.id,
            "product_id":    a.product_id,
            "variant_id":    a.variant_id,
            "variant_name":  str(a.variant),
            "sku":           a.variant.sku,
            "quantity":      a.quantity,
            "stock_left":    stock_left,
            "has_auto_rule": has_auto_rule,
        })

    orders = sorted(order_map.values(), key=lambda x: x["order_id"], reverse=True)
    return JsonResponse({"success": True, "orders": orders})
